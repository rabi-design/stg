# -*- coding: utf-8 -*-
import openpyxl

file = "重複チェックファイル.xlsx"
book1 = openpyxl.load_workbook(file)
sheet1 = book1['シート1']

ro = 2
list1 = []
while True:
    list1.append(sheet1.cell(ro, 2).value)
    ro += 1
    if sheet1.cell(ro, 2).value is None:
        if sheet1.cell(ro + 1, 2).value is None:
            if sheet1.cell(ro + 2, 2).value is None:
                break
with open('list_other.py', 'w') as f:
    f.write("urls=[\n")
    for d in list1:
        f.write("'%s',\n" % d)
    f.write("]")
